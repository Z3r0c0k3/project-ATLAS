from __future__ import annotations

import re
import shutil
import subprocess
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .integrity import canonical_sha256, file_sha256
from .naming import normalize_filename


class WorkbookParseError(ValueError):
    pass


LEDGER_HEADERS = {"날짜", "내용", "수입", "지출"}


def _label(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def _parse_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = _label(value)
    if not text:
        return None
    for pattern in (
        "%Y.%m.%d %H:%M:%S",
        "%Y.%m.%d %H:%M",
        "%Y. %m. %d %H:%M:%S",
        "%Y. %m. %d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y.%m.%d",
        "%Y. %m. %d",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def _money(value: Any) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(round(value))
    text = _label(value).replace(",", "").replace("원", "").replace("₩", "")
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    try:
        number = int(round(float(text)))
    except ValueError as exc:
        raise WorkbookParseError(f"금액을 숫자로 읽을 수 없습니다: {value}") from exc
    return -number if negative else number


def _find_header(sheet, required: set[str]) -> tuple[int | None, dict[str, int]]:
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 60)):
        columns: dict[str, int] = {}
        for cell in row:
            text = _label(cell.value)
            if text:
                columns.setdefault(text, cell.column)
        if required.issubset(columns):
            return row[0].row, columns
    return None, {}


def _open(path: Path):
    if path.suffix.lower() == ".xls":
        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        if not soffice:
            raise WorkbookParseError(".xls 파일을 읽으려면 LibreOffice가 필요합니다.")
        try:
            with tempfile.TemporaryDirectory(prefix="atlas-xls-") as temporary:
                subprocess.run(
                    [
                        soffice,
                        f"-env:UserInstallation={Path(temporary, 'profile').as_uri()}",
                        "--headless",
                        "--convert-to",
                        "xlsx",
                        "--outdir",
                        temporary,
                        str(path),
                    ],
                    check=True,
                    capture_output=True,
                    timeout=60,
                )
                converted = next(Path(temporary).glob("*.xlsx"), None)
                if not converted:
                    raise WorkbookParseError(f"Excel 파일을 변환할 수 없습니다: {normalize_filename(path.name)}")
                return load_workbook(converted, data_only=True, read_only=False)
        except (subprocess.SubprocessError, OSError) as exc:
            raise WorkbookParseError(f"Excel 파일을 변환할 수 없습니다: {normalize_filename(path.name)}") from exc
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise WorkbookParseError(".xls, .xlsx 또는 .xlsm 파일만 가져올 수 있습니다.")
    try:
        return load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        raise WorkbookParseError(f"Excel 파일을 열 수 없습니다: {normalize_filename(path.name)}") from exc


def detect_workbook(path: Path) -> dict:
    workbook = _open(path)
    matches: list[dict] = []
    for sheet in workbook.worksheets:
        ledger_row, ledger_columns = _find_header(sheet, LEDGER_HEADERS)
        if ledger_row:
            matches.append({"kind": "aegis_ledger", "sheet_name": sheet.title, "header_row": ledger_row, "columns": ledger_columns})
    if not matches:
        raise WorkbookParseError("지원하는 Aegis 장부 헤더를 찾지 못했습니다.")
    kinds = {item["kind"] for item in matches}
    return {
        "kind": next(iter(kinds)) if len(kinds) == 1 else "mixed",
        "sha256": file_sha256(path),
        "sheets": matches,
    }


def _category(description: str, method: str) -> str:
    text = f"{description} {method}".lower()
    rules = (
        ("회비", "회비"),
        ("지원금", "지원금"),
        ("이자", "이자"),
        ("도메인", "IT/서버"),
        ("서버", "IT/서버"),
        ("간식", "복리후생"),
        ("총회", "행사"),
        ("mt", "행사"),
        ("마라톤", "행사/상품"),
        ("경품", "행사/상품"),
        ("상품", "행사/상품"),
        ("비품", "비품"),
        ("홍보", "홍보"),
    )
    for keyword, category in rules:
        if keyword in text:
            return category
    return "미분류"


def _counterparty(detail: str) -> str | None:
    for pattern in (r"결제처\s*:\s*([^,]+)", r"입금자(?:명)?\s*:\s*([^,]+)", r"지출자\s*:\s*([^,]+)"):
        match = re.search(pattern, detail)
        if match:
            return match.group(1).strip()
    return None


def parse_aegis_ledger(
    path: Path,
    *,
    period: str,
    opening_balance: int = 0,
    organization_id: str = "aegis",
    account_id: str = "primary",
    sheet_name: str | None = None,
) -> dict:
    workbook = _open(path)
    selected = None
    header_row = None
    columns: dict[str, int] = {}
    for sheet in workbook.worksheets:
        row, found = _find_header(sheet, LEDGER_HEADERS)
        if row and (sheet_name is None or sheet.title == sheet_name):
            selected, header_row, columns = sheet, row, found
            break
    if selected is None or header_row is None:
        raise WorkbookParseError("Aegis 장부 헤더(날짜/내용/수입/지출)를 찾지 못했습니다.")

    running_balance = opening_balance
    transactions: list[dict] = []
    empty_streak = 0
    for row_number in range(header_row + 1, selected.max_row + 1):
        tx_date = _parse_date(selected.cell(row_number, columns["날짜"]).value)
        description = _label(selected.cell(row_number, columns["내용"]).value)
        income = _money(selected.cell(row_number, columns["수입"]).value)
        expense = _money(selected.cell(row_number, columns["지출"]).value)
        if not tx_date and not description and not income and not expense:
            empty_streak += 1
            if transactions and empty_streak >= 50:
                break
            continue
        empty_streak = 0
        if not tx_date:
            continue

        number_column = columns.get("NO") or columns.get("번호")
        raw_number = selected.cell(row_number, number_column).value if number_column else None
        number = _money(raw_number) or len(transactions) + 1
        running_balance += income - expense
        raw_balance = selected.cell(row_number, columns.get("잔액", 0)).value if columns.get("잔액") else None
        if isinstance(raw_balance, (int, float)):
            running_balance = _money(raw_balance)
        method = _label(selected.cell(row_number, columns.get("처리방식", 0)).value) if columns.get("처리방식") else ""
        detail = _label(selected.cell(row_number, columns.get("상세정보", 0)).value) if columns.get("상세정보") else ""
        note = " / ".join(value for value in (method, detail) if value)
        source_row = f"{selected.title}!{row_number}"
        transaction = {
            "organization_id": organization_id,
            "account_id": account_id,
            "period": period,
            "number": number,
            "date": tx_date.date().isoformat(),
            "description": description,
            "income": income,
            "expense": expense,
            "balance": running_balance,
            "category": _category(description, method),
            "counterparty": _counterparty(detail),
            "processing_method": method,
            "details": detail,
            "note": note,
            "source_row": source_row,
        }
        transaction["source_row_hash"] = canonical_sha256(transaction)
        transaction["transaction_id"] = f"tx_{transaction['source_row_hash'][:16]}"
        transactions.append(transaction)

    if not transactions:
        raise WorkbookParseError("장부에서 유효한 거래 행을 찾지 못했습니다.")
    return {
        "kind": "aegis_ledger",
        "filename": normalize_filename(path.name),
        "sha256": file_sha256(path),
        "sheet_name": selected.title,
        "header_row": header_row,
        "opening_balance": opening_balance,
        "closing_balance": transactions[-1]["balance"],
        "period_start": transactions[0]["date"],
        "period_end": transactions[-1]["date"],
        "total_income": sum(item["income"] for item in transactions),
        "total_expense": sum(item["expense"] for item in transactions),
        "transaction_count": len(transactions),
        "transactions": transactions,
    }


def workbook_preview(path: Path) -> dict:
    detected = detect_workbook(path)
    kind = detected["kind"]
    if kind == "aegis_ledger":
        parsed = parse_aegis_ledger(path, period="미지정")
        return {
            **detected,
            "summary": {
                key: parsed[key]
                for key in ("sheet_name", "transaction_count", "period_start", "period_end", "total_income", "total_expense", "closing_balance")
            },
            "preview": parsed["transactions"][:10],
        }
    return detected
