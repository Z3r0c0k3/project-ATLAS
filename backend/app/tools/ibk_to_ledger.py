from __future__ import annotations

import argparse
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..services.workbooks import parse_ibk_bank_pdf

TRANSACTION_TYPE_MAP = {
    "스마트뱅킹": "계좌이체",
    "타행이체": "계좌이체",
    "펌뱅킹": "계좌이체",
    "이자": "이자입금",
}


def _format_amount(amount: int) -> str:
    return f"{amount:,}"


def _date_from_cell(value: str) -> datetime:
    return datetime.fromisoformat(value[:10])


def ibk_pdf_to_aegis_ledger(
    pdf_path: str | Path,
    output_path: str | Path | None = None,
    club_name: str = "Aegis",
    semester: str = "",
    opening_balance: int | None = None,
) -> Path:
    pdf_path = Path(pdf_path)
    parsed = parse_ibk_bank_pdf(pdf_path)

    if opening_balance is None:
        opening_balance = parsed["opening_balance"]

    if output_path is None:
        account_holder = parsed.get("account_holder", "회비계좌")
        output_path = pdf_path.with_name(f"{pdf_path.stem}.xlsx")
    output_path = Path(output_path)

    wb = Workbook()
    ws = wb.active
    ws.title = semester or "회비입금계좌"
    ws.sheet_properties.tabColor = "2F5496"

    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="2F5496")
    body_font = Font(name="맑은 고딕", size=10)
    number_font = Font(name="맑은 고딕", size=10)
    money_fill_in = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    money_fill_out = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    title_font = Font(name="맑은 고딕", size=14, bold=True, color="1F3864")

    ws.merge_cells("B1:I1")
    ws["B1"] = f"{club_name} 회비입금계좌 회계장부"
    ws["B1"].font = title_font
    ws["B1"].alignment = center_align

    account_info_text = f"계좌번호: {parsed.get('account_number', '')} | 예금주: {parsed.get('account_holder', '')} | 조회기간: {parsed.get('query_start', '')} ~ {parsed.get('query_end', '')}"
    ws.merge_cells("B2:I2")
    ws["B2"] = account_info_text
    ws["B2"].font = Font(name="맑은 고딕", size=9, color="666666")
    ws["B2"].alignment = center_align

    ws.merge_cells("B3:I3")
    ws["B3"] = (
        f"이전잔액: {_format_amount(opening_balance)}원 | "
        f"총입금: {_format_amount(parsed['total_inflow'])}원 | "
        f"총출금: {_format_amount(parsed['total_outflow'])}원 | "
        f"최종잔액: {_format_amount(parsed['closing_balance'])}원"
    )
    ws["B3"].font = Font(name="맑은 고딕", size=10, color="444444")
    ws["B3"].alignment = center_align

    headers = ["NO", "날짜", "내용", "수입", "지출", "잔액", "처리방식", "상세정보"]
    col_letters = ["B", "C", "D", "E", "F", "G", "H", "I"]
    col_widths = [6, 14, 42, 15, 15, 15, 12, 50]

    header_row = 5
    for col_idx, (letter, header, width) in enumerate(zip(col_letters, headers, col_widths)):
        cell = ws[f"{letter}{header_row}"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws.column_dimensions[letter].width = width

    transactions = parsed["transactions"]
    running_balance = opening_balance

    for idx, tx in enumerate(transactions, start=1):
        row = header_row + idx
        date_val = _date_from_cell(tx["date"])
        deposit = tx["deposit"]
        withdrawal = tx["withdrawal"]
        running_balance = running_balance + deposit - withdrawal

        tx_type = "".join(tx.get("transaction_type", "").split())
        processing_method = TRANSACTION_TYPE_MAP.get(tx_type, "")

        counterparty = tx.get("counterparty_name", "") or tx.get("description", "")
        message = tx.get("transfer_message", "")

        if deposit > 0:
            description = f"{counterparty}{' - ' + message if message else ''}"
            details = f"입금자: {counterparty}"
            if message and message not in details:
                details += f", 비고: {message}"
        else:
            description = f"{counterparty}{' - ' + message if message else ''}"
            details = f"이체대상: {counterparty}"
            if message and message not in details:
                details += f", 비고: {message}"
            if "회비환불" in tx.get("description", ""):
                description = tx.get("description", description)
                details = f"출금: {counterparty}"

        if tx_type == "이자":
            description = "통장 이자"
            details = ""
            if message:
                details = message

        row_data = {
            "B": idx,
            "C": date_val,
            "D": description.strip(),
            "E": deposit if deposit > 0 else None,
            "F": withdrawal if withdrawal > 0 else None,
            "G": running_balance,
            "H": processing_method,
            "I": details.strip() if details.strip() else None,
        }

        for letter, value in row_data.items():
            cell = ws[f"{letter}{row}"]
            cell.value = value
            cell.font = body_font
            cell.border = thin_border

            if letter in ("B", "C", "H"):
                cell.alignment = center_align
                if letter == "B":
                    cell.font = number_font
            elif letter in ("E", "F", "G"):
                cell.alignment = right_align
                cell.number_format = "#,##0"
                if letter == "E" and deposit > 0:
                    cell.fill = money_fill_in
                elif letter == "F" and withdrawal > 0:
                    cell.fill = money_fill_out
            else:
                cell.alignment = left_align

    last_data_row = header_row + len(transactions)
    ws.merge_cells(f"B{last_data_row + 1}:D{last_data_row + 1}")
    summary_cell = ws[f"B{last_data_row + 1}"]
    summary_cell.value = "합계"
    summary_cell.font = Font(name="맑은 고딕", size=10, bold=True, color="2F5496")
    summary_cell.alignment = center_align
    summary_cell.border = thin_border

    for col_letter in ("E", "F"):
        cell = ws[f"{col_letter}{last_data_row + 1}"]
        cell.value = f"=SUM({col_letter}{header_row + 1}:{col_letter}{last_data_row})"
        cell.font = Font(name="맑은 고딕", size=10, bold=True)
        cell.alignment = right_align
        cell.number_format = "#,##0"
        cell.border = thin_border

    ws[f"G{last_data_row + 1}"].border = thin_border
    ws[f"H{last_data_row + 1}"].border = thin_border
    ws[f"I{last_data_row + 1}"].border = thin_border

    ws.auto_filter.ref = f"B{header_row}:I{last_data_row}"
    ws.freeze_panes = f"B{header_row + 1}"

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="IBK 기업은행 PDF 거래내역 → Aegis 회계장부 Excel 변환기")
    parser.add_argument("pdf_path", type=str, help="IBK 기업은행 거래내역 PDF 파일 경로")
    parser.add_argument(
        "-o", "--output", type=str, default=None, help="출력 Excel 파일 경로 (기본값: PDF와 동일 경로에 .xlsx)"
    )
    parser.add_argument("--club-name", type=str, default="Aegis", help="동아리명 (기본값: Aegis)")
    parser.add_argument("--semester", type=str, default="", help="학기명 (시트명으로 사용)")
    parser.add_argument("--opening-balance", type=int, default=None, help="이전 잔액 (기본값: PDF에서 자동 계산)")
    args = parser.parse_args()

    output = ibk_pdf_to_aegis_ledger(
        pdf_path=args.pdf_path,
        output_path=args.output,
        club_name=args.club_name,
        semester=args.semester,
        opening_balance=args.opening_balance,
    )
    print(f"변환 완료: {output}")


if __name__ == "__main__":
    main()
